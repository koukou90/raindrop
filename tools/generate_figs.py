from __future__ import annotations

import pickle
import sys
from pathlib import Path
from typing import Dict, List, Tuple

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.colors import TwoSlopeNorm
from matplotlib.gridspec import GridSpec
from matplotlib.patches import Patch, Rectangle
from openpyxl import load_workbook


# =========================
# 基础路径与常量
# =========================
ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "out"
FIG_DIR = OUT_DIR / "figs"

if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from utils.metrics import mae as metric_mae
from utils.metrics import mse as metric_mse
from utils.metrics import nse as metric_nse
from utils.metrics import peak_mae as metric_peak_mae
from utils.metrics import rmse as metric_rmse
from utils.metrics import compute_all_metrics

SITE_FULL = ["W2127_Haichaoba", "W2128_Haichaoyinsi", "W2129_buligou"]
SITE_CODE = {
    "W2127_Haichaoba": "W2127",
    "W2128_Haichaoyinsi": "W2128",
    "W2129_buligou": "W2129",
}
SITE_CODE_INV = {v: k for k, v in SITE_CODE.items()}

PROPOSED_CANON = "TripleStreamV3_self_distill"
METRIC_COLS = ["RMSE", "MAE", "MSE", "NSE", "Peak-MAE"]
EXCLUDED_MODEL_CANONS = {"BaselineTiDE"}


def _set_global_style() -> None:
    plt.rcParams.update(
        {
            "font.family": "Times New Roman",
            "axes.labelsize": 12,
            "axes.titlesize": 12,
            "xtick.labelsize": 11,
            "ytick.labelsize": 11,
            "legend.fontsize": 10,
            "savefig.bbox": "tight",
        }
    )


def _to_float(x):
    try:
        return float(x)
    except Exception:
        return np.nan


# =========================
# 模型名规范化与展示名
# =========================
CANONICAL_MODELS = [
    "TripleStreamV3_self_distill",
    "BaselineLightGBM",
    "BaselineMLP",
    "BaselineLSTM",
    "BaselineDLinear",
    "BaselinePatchTST",
    "BaselineTiDE",
    "BaselineCNNTransformer",
    "BaselineiTransformer",
    "AblateV3NoPersistence",
    "AblateV3NoMixGate",
    "AblateV3NoBinAware",
    "AblateV3NoAux",
    "AblateV3NoStreamGate",
]

MODEL_ALIASES = {
    "TripleStreamV3_self_distill": ["TripleStreamV3"],
    "BaselineLightGBM": ["LightGBM"],
    "BaselineMLP": ["MLP"],
    "BaselineLSTM": ["LSTM"],
    "BaselineDLinear": ["BaselineLinear", "DLinear"],
    "BaselinePatchTST": ["PatchTST"],
    "BaselineTiDE": ["TiDE"],
    "BaselineCNNTransformer": ["CNN-Transformer"],
    "BaselineiTransformer": ["iTransformer"],
    "AblateV3NoPersistence": ["NoPersistence"],
    "AblateV3NoMixGate": ["NoMixGate"],
    "AblateV3NoBinAware": ["NoBinAware"],
    "AblateV3NoAux": ["NoAux"],
    "AblateV3NoStreamGate": ["NoStreamGate"],
}

CANONICAL_NAME_MAP: Dict[str, str] = {}
for _canon in CANONICAL_MODELS:
    CANONICAL_NAME_MAP[_canon] = _canon
    for _alias in MODEL_ALIASES.get(_canon, []):
        CANONICAL_NAME_MAP[_alias] = _canon


def canonical_model_name(name: str) -> str:
    if name is None:
        return ""
    n = str(name).strip()
    return CANONICAL_NAME_MAP.get(n, n)


def model_display_name(canon: str) -> str:
    mapping = {
        "TripleStreamV3_self_distill": "Ours",
        "BaselineLightGBM": "LightGBM",
        "BaselineMLP": "MLP",
        "BaselineLSTM": "LSTM",
        "BaselineDLinear": "DLinear",
        "BaselinePatchTST": "PatchTST",
        "BaselineTiDE": "TiDE",
        "BaselineCNNTransformer": "CNN-Transformer",
        "BaselineiTransformer": "iTransformer",
        "AblateV3NoPersistence": "NoPersistence",
        "AblateV3NoMixGate": "NoMixGate",
        "AblateV3NoBinAware": "NoBinAware",
        "AblateV3NoAux": "NoAux",
        "AblateV3NoStreamGate": "NoStreamGate",
    }
    return mapping.get(canon, canon)


def model_run_prefixes(canon: str) -> List[str]:
    if canon == "TripleStreamV3_self_distill":
        return ["TripleStreamV3"]
    if canon == "BaselineDLinear":
        return ["BaselineDLinear", "BaselineLinear"]
    return [canon]


# =========================
# 读取整体对比表与消融表
# =========================
def read_overall_site_table(site_full: str) -> pd.DataFrame:
    xlsx = OUT_DIR / site_full / f"metrics_comparison_template_{site_full}.xlsx"
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Overall"]

    headers = [ws.cell(2, c).value for c in range(1, 7)]
    rows: List[List[object]] = []
    r = 3
    while True:
        model = ws.cell(r, 1).value
        if model is None:
            break
        rows.append([ws.cell(r, c).value for c in range(1, 7)])
        r += 1

    df = pd.DataFrame(rows, columns=headers)
    df["Site"] = SITE_CODE[site_full]
    df["ModelCanon"] = df["Model"].map(canonical_model_name)
    df = df[~df["ModelCanon"].isin(EXCLUDED_MODEL_CANONS)].reset_index(drop=True)
    df["ModelDisplay"] = df["ModelCanon"].map(model_display_name)
    for col in METRIC_COLS:
        df[col] = df[col].map(_to_float)
    return df


def read_ablation_site_table(site_full: str) -> pd.DataFrame:
    pattern = f"metrics_ablation_{site_full}.xlsx"
    matches = list((OUT_DIR / site_full).rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Cannot find {pattern}")

    wb = load_workbook(matches[0], data_only=True)
    ws = wb["Ablation"]
    headers = [ws.cell(2, c).value for c in range(1, 7)]
    rows: List[List[object]] = []
    r = 3
    while True:
        model = ws.cell(r, 1).value
        if model is None:
            break
        rows.append([ws.cell(r, c).value for c in range(1, 7)])
        r += 1

    df = pd.DataFrame(rows, columns=headers)
    df["Site"] = SITE_CODE[site_full]
    df["ModelCanon"] = df["Model"].map(canonical_model_name)
    df["ModelDisplay"] = df["ModelCanon"].map(model_display_name)
    for col in METRIC_COLS:
        df[col] = df[col].map(_to_float)
    return df


def select_best_baseline(overall_df: pd.DataFrame) -> str:
    tmp = overall_df[overall_df["ModelCanon"] != PROPOSED_CANON].copy()
    return tmp.sort_values("RMSE").iloc[0]["ModelCanon"]


# =========================
# 事件级数据读取（用于 Fig5/Fig6）
# =========================
def find_run_dir(site_full: str, model_canon: str) -> Path:
    site_dir = OUT_DIR / site_full
    candidates: List[Path] = []
    for prefix in model_run_prefixes(model_canon):
        candidates.extend([p for p in site_dir.iterdir() if p.is_dir() and p.name.startswith(prefix)])
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    for folder in candidates:
        pred_pkl = folder / "timeline_predictions_by_event.pkl"
        lab_pkl = folder / "timeline_labels_by_event.pkl"
        if pred_pkl.exists() and lab_pkl.exists():
            return folder
    raise FileNotFoundError(f"No event-level output found: {site_full} / {model_canon}")


def load_event_series(run_dir: Path) -> Tuple[List[np.ndarray], List[np.ndarray]]:
    pred = pickle.load(open(run_dir / "timeline_predictions_by_event.pkl", "rb"))
    lab = pickle.load(open(run_dir / "timeline_labels_by_event.pkl", "rb"))
    if isinstance(pred, dict):
        keys = list(pred.keys())
        pred_list = [np.asarray(pred[k]).flatten() for k in keys]
        lab_list = [np.asarray(lab[k]).flatten() for k in keys]
        return pred_list, lab_list
    pred_list = [np.asarray(x).flatten() for x in pred]
    lab_list = [np.asarray(x).flatten() for x in lab]
    return pred_list, lab_list


# =========================
# 指标计算
# =========================
def compute_overall(y_true: np.ndarray, y_pred: np.ndarray) -> Dict[str, float]:
    return {
        "RMSE": metric_rmse(y_true, y_pred),
        "MAE": metric_mae(y_true, y_pred),
        "MSE": metric_mse(y_true, y_pred),
        "NSE": metric_nse(y_true, y_pred),
        "Peak-MAE": metric_peak_mae(y_true, y_pred),
    }


# =========================
# 构建 Fig5/Fig6/Fig7 数据
# =========================
def _intensity_group(peak: float) -> str:
    if peak < 2.0:
        return "Weak"
    if peak < 10.0:
        return "Moderate"
    return "Strong"


def _duration_group(length: int) -> str:
    if length < 60:
        return "Short"
    if length < 180:
        return "Medium"
    return "Long"


def build_scenario_delta_table(site_best_baseline: Dict[str, str]) -> pd.DataFrame:
    containers: Dict[Tuple[str, str, str], Dict[str, List[np.ndarray]]] = {}
    for site_full in SITE_FULL:
        site = SITE_CODE[site_full]
        baseline = site_best_baseline[site]
        run_main = find_run_dir(site_full, PROPOSED_CANON)
        run_base = find_run_dir(site_full, baseline)
        p_main, y_main = load_event_series(run_main)
        p_base, y_base = load_event_series(run_base)
        n_events = min(len(p_main), len(p_base), len(y_main), len(y_base))
        for i in range(n_events):
            y = np.asarray(y_main[i]).flatten()
            p_m = np.asarray(p_main[i]).flatten()
            p_b = np.asarray(p_base[i]).flatten()
            n = min(len(y), len(p_m), len(p_b))
            if n <= 0:
                continue
            y, p_m, p_b = y[:n], p_m[:n], p_b[:n]
            group_i = _intensity_group(float(np.max(y)))
            group_d = _duration_group(int(len(y)))
            for stype, sgroup in [("Intensity", group_i), ("Duration", group_d)]:
                key = (site, stype, sgroup)
                if key not in containers:
                    containers[key] = {"y": [], "m": [], "b": []}
                containers[key]["y"].append(y)
                containers[key]["m"].append(p_m)
                containers[key]["b"].append(p_b)

    rows = []
    for (site, stype, sgroup), pack in containers.items():
        y = np.concatenate(pack["y"])
        m = np.concatenate(pack["m"])
        b = np.concatenate(pack["b"])
        rmse_m = metric_rmse(y, m)
        rmse_b = metric_rmse(y, b)
        rows.append(
            {
                "Site": site,
                "ScenarioType": stype,
                "Scenario": sgroup,
                "N_events": len(pack["y"]),
                "N_points": len(y),
                "Delta_RMSE_pct": (rmse_b - rmse_m) / rmse_b * 100.0 if rmse_b > 0 else np.nan,
            }
        )
    return pd.DataFrame(rows)


def build_case_samples(
    site_best_baseline: Dict[str, str],
    case_index_by_site: Dict[str, int] | None = None,
) -> List[Dict[str, object]]:
    cases: List[Dict[str, object]] = []
    for site_full in SITE_FULL:
        site = SITE_CODE[site_full]
        baseline = site_best_baseline[site]
        run_main = find_run_dir(site_full, PROPOSED_CANON)
        run_base = find_run_dir(site_full, baseline)
        p_main, y_main = load_event_series(run_main)
        p_base, y_base = load_event_series(run_base)
        n_events = min(len(p_main), len(p_base), len(y_main), len(y_base))
        if n_events <= 0:
            continue

        if case_index_by_site and site in case_index_by_site:
            # User-facing case index is 1-based; 0 is also accepted for first event.
            raw_idx = int(case_index_by_site[site])
            idx = raw_idx - 1 if raw_idx >= 1 else raw_idx
            idx = max(0, min(idx, n_events - 1))
        else:
            peaks = [float(np.max(np.asarray(y_main[i]).flatten())) for i in range(n_events)]
            idx = int(np.argmax(peaks))
        y = np.asarray(y_main[idx]).flatten()
        m = np.asarray(p_main[idx]).flatten()
        b = np.asarray(p_base[idx]).flatten()
        n = min(len(y), len(m), len(b))
        y, m, b = y[:n], m[:n], b[:n]
        rmse_m = metric_rmse(y, m)
        rmse_b = metric_rmse(y, b)
        mae_m = metric_mae(y, m)
        mae_b = metric_mae(y, b)
        cases.append(
            {
                "site": site,
                "case_id": f"{site}_Event{idx+1}",
                "baseline_display": model_display_name(baseline),
                "y_true": y,
                "y_main": m,
                "y_base": b,
                "delta_rmse_pct": (rmse_b - rmse_m) / rmse_b * 100.0 if rmse_b > 0 else np.nan,
                "delta_mae_pct": (mae_b - mae_m) / mae_b * 100.0 if mae_b > 0 else np.nan,
                "rmse_main": rmse_m,
                "rmse_base": rmse_b,
                "mae_main": mae_m,
                "mae_base": mae_b,
            }
        )
    return cases


def build_ablation_delta_table(ablation_dfs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: List[pd.DataFrame] = []
    for site, df in ablation_dfs.items():
        main_row = df[df["ModelCanon"] == PROPOSED_CANON].iloc[0]
        tmp = df.copy()
        for metric in ["RMSE", "MAE", "MSE"]:
            main_val = float(main_row[metric])
            col = f"Delta_{metric}_vs_Proposed_pct"
            tmp[col] = (tmp[metric] - main_val) / main_val * 100.0 if main_val > 0 else np.nan
        # Use percentage points for NSE to avoid unstable relative percentages near zero.
        tmp["Delta_NSE_vs_Proposed_pp"] = (float(main_row["NSE"]) - tmp["NSE"]) * 100.0
        rows.append(tmp)
    out = pd.concat(rows, ignore_index=True)
    out = out[
        [
            "Site",
            "ModelCanon",
            "ModelDisplay",
            "RMSE",
            "MAE",
            "MSE",
            "NSE",
            "Delta_RMSE_vs_Proposed_pct",
            "Delta_MAE_vs_Proposed_pct",
            "Delta_MSE_vs_Proposed_pct",
            "Delta_NSE_vs_Proposed_pp",
        ]
    ]
    return out


# =========================
# 公共数据准备（供手动执行代码块调用）
# =========================
def _prepare_data():
    overall_dfs: Dict[str, pd.DataFrame] = {}
    ablation_dfs: Dict[str, pd.DataFrame] = {}
    site_best_baseline: Dict[str, str] = {}

    for site_full in SITE_FULL:
        site = SITE_CODE[site_full]
        odf = read_overall_site_table(site_full)
        adf = read_ablation_site_table(site_full)
        overall_dfs[site] = odf
        ablation_dfs[site] = adf
        site_best_baseline[site] = select_best_baseline(odf)

    scenario_df = build_scenario_delta_table(site_best_baseline)
    case_samples = build_case_samples(site_best_baseline)
    ablation_delta_df = build_ablation_delta_table(ablation_dfs)
    return overall_dfs, site_best_baseline, scenario_df, case_samples, ablation_delta_df


# =========================
# Figure 3
# =========================
def plot_figure3_sitewise_multimodel_curves(
    overall_dfs: Dict[str, pd.DataFrame],
    out_path: Path | None = None,
) -> Path:
    """
    Figure 3:
    3x1 site-wise prediction curve comparison.
    Each subplot: observation + proposed model + all baselines (TiDE excluded).
    """
    cfg = {
        "figsize": (12, 6),
        "hspace": 0.32,
        "title_size": 10.0,
        "label_size": 12.5,
        "tick_size": 10.8,
        "legend_size": 8,
        "metrics_size": 10,
        "color_truth": "#d62728",
        "color_main": "#1f77b4",
        "color_boundary": "#9e9e9e",
        "line_truth": 1.25,
        "line_main": 1.35,
        "line_base": 1.05,
    }

    fig, axes = plt.subplots(3, 1, figsize=cfg["figsize"])
    fig.subplots_adjust(hspace=cfg["hspace"], top=0.93)

    panel_tags = ["(a)", "(b)", "(c)"]
    baseline_cmap = plt.get_cmap("tab20")

    for i, site_full in enumerate(SITE_FULL):
        ax = axes[i]
        site = SITE_CODE[site_full]
        df = overall_dfs[site].copy()

        baseline_models = df[
            (df["ModelCanon"] != PROPOSED_CANON) & (df["ModelCanon"].astype(str).str.startswith("Baseline"))
        ].copy()
        baseline_models = baseline_models.sort_values("RMSE")
        baseline_list = baseline_models["ModelCanon"].dropna().tolist()
        model_order = [PROPOSED_CANON] + baseline_list

        run_main = find_run_dir(site_full, PROPOSED_CANON)
        _, y_main_list = load_event_series(run_main)
        if len(y_main_list) == 0:
            continue

        pred_events_map: Dict[str, List[np.ndarray]] = {}
        available_models: List[str] = []

        for model_canon in model_order:
            try:
                run_dir = find_run_dir(site_full, model_canon)
            except FileNotFoundError:
                continue
            pred_list, _ = load_event_series(run_dir)
            if len(pred_list) == 0:
                continue
            pred_events_map[model_canon] = [np.asarray(x).flatten() for x in pred_list]
            available_models.append(model_canon)

        if PROPOSED_CANON not in available_models:
            continue

        n_events = min([len(y_main_list)] + [len(pred_events_map[m]) for m in available_models])
        if n_events <= 0:
            continue

        true_concat: List[float] = []
        pred_concat: Dict[str, List[float]] = {m: [] for m in available_models}
        event_boundaries = [0]

        for evt_idx in range(n_events):
            y_evt = np.asarray(y_main_list[evt_idx]).flatten()
            evt_len = min([len(y_evt)] + [len(pred_events_map[m][evt_idx]) for m in available_models])
            if evt_len <= 0:
                continue

            true_concat.extend(y_evt[:evt_len].tolist())
            for m in available_models:
                pred_concat[m].extend(pred_events_map[m][evt_idx][:evt_len].tolist())
            event_boundaries.append(len(true_concat))

        if len(true_concat) == 0:
            continue

        y_true = np.asarray(true_concat, dtype=float)
        x = np.arange(len(y_true))
        ax.plot(
            x,
            y_true,
            color=cfg["color_truth"],
            lw=cfg["line_truth"],
            label="Observation",
            alpha=0.9,
            zorder=6,
        )

        for k, model_canon in enumerate(model_order):
            if model_canon not in pred_concat:
                continue
            y_pred = np.asarray(pred_concat[model_canon], dtype=float)
            if len(y_pred) != len(y_true):
                continue
            display = model_display_name(model_canon)
            if model_canon == PROPOSED_CANON:
                ax.plot(
                    x,
                    y_pred,
                    color=cfg["color_main"],
                    lw=cfg["line_main"],
                    label=display,
                    alpha=0.95,
                    zorder=7,
                )
            else:
                ax.plot(
                    x,
                    y_pred,
                    color=baseline_cmap((k - 1) % 20),
                    lw=cfg["line_base"],
                    ls="--",
                    alpha=0.9,
                    label=display,
                    zorder=3,
                )

        for boundary in event_boundaries[1:-1]:
            ax.axvline(
                x=boundary,
                color=cfg["color_boundary"],
                linestyle="--",
                linewidth=0.75,
                alpha=0.35,
                zorder=1,
            )

        main_pred = np.asarray(pred_concat[PROPOSED_CANON], dtype=float)
        metrics = compute_all_metrics(y_true, main_pred, threshold=10.0)
        peak_mae_str = f"{metrics['Peak_MAE']:.3f}" if not np.isnan(metrics["Peak_MAE"]) else "N/A"
        metrics_text = (
            f"RMSE={metrics['RMSE']:.3f} | "
            f"MAE={metrics['MAE']:.3f} | "
            f"Peak-MAE={peak_mae_str}"
        )
        ax.text(
            0.62,
            0.98,
            metrics_text,
            transform=ax.transAxes,
            ha="left",
            va="top",
            fontsize=cfg["metrics_size"],
        )

        ax.set_title(
            f"{panel_tags[i]} {site} (Precipitation Events: {len(event_boundaries)-1})",
            loc="left",
            fontsize=cfg["title_size"],
            fontweight="bold",
            pad=6,
        )
        ax.set_ylabel("Rain rate (mm/h)", fontsize=cfg["label_size"], fontweight="bold")
        ax.tick_params(axis="both", labelsize=cfg["tick_size"], direction="out", length=3, width=0.8)
        if i == len(SITE_FULL) - 1:
            ax.set_xlabel("Timeline Index (Concatenated by Events)", fontsize=cfg["label_size"], fontweight="bold")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)

    handles, labels = axes[0].get_legend_handles_labels()
    fig.legend(
        handles,
        labels,
        loc="upper center",
        ncol=9,
        frameon=False,
        fancybox=False,
        framealpha=0.92,
        fontsize=cfg["legend_size"],
        bbox_to_anchor=(0.5, 0.999),
    )

    out = out_path or (FIG_DIR / "Sitewise_predictions_multimodel.tiff")
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=300)
    plt.close(fig)
    return out


# 手动执行（Figure 3）
# _set_global_style()
# FIG_DIR.mkdir(parents=True, exist_ok=True)
# overall_dfs, site_best_baseline, scenario_df, case_samples, ablation_delta_df = _prepare_data()
# out = plot_figure3_sitewise_multimodel_curves(overall_dfs)
# print(f"[saved] {out}")


# =========================
# Figure 4
# =========================
def plot_figure4_overall_cross_site(
    overall_dfs: Dict[str, pd.DataFrame],
    site_best_baseline: Dict[str, str],
    out_path: Path | None = None,
) -> Path:
    """
    Figure 4:
    2x3 multi-panel leaderboard:
    top row  : RMSE ranking at W2127/W2128/W2129
    bottom row: NSE ranking at W2127/W2128/W2129
    """
    # ---- Figure 4 config (可单独调参) ----
    cfg = {
        "figsize": (16.8, 8.2),
        "hspace": 0.16,
        "wspace": 0.24,
        "color_other": "#BFC8D6",
        "color_best": "#B56A2C",
        "color_main": "#2C7FB8",
        "panel_face": "#F7F9FC",
        "stripe_color": "#EDF2F7",
        "grid_color": "#A5B4C6",
        "title_box_face": "#E6ECF4",
        "value_box_edge": "#D2DAE5",
        "spine_color": "#CDD6E2",
        "bar_height": 0.62,
        "annot_size": 10,
        "ytick_size": 9.2,
        "xtick_size": 11,
        "label_size": 12,
        "panel_size": 11,
        "note_size": 10,
    }

    site_order = ["W2127", "W2128", "W2129"]
    metrics = [("RMSE", True), ("NSE", False)]  # (metric, lower_is_better)
    panel_tags = ["(a)", "(b)", "(c)", "(d)", "(e)", "(f)"]

    rmse_vals_all = pd.concat([overall_dfs[s]["RMSE"] for s in site_order], ignore_index=True).dropna().values
    nse_vals_all = pd.concat([overall_dfs[s]["NSE"] for s in site_order], ignore_index=True).dropna().values
    rmse_span = max(float(np.max(rmse_vals_all)) - float(np.min(rmse_vals_all)), 1e-6)
    nse_span = max(float(np.max(nse_vals_all)) - float(np.min(nse_vals_all)), 1e-6)
    row_xlim = {
        "RMSE": (
            max(0.0, float(np.min(rmse_vals_all)) - rmse_span * 0.08),
            float(np.max(rmse_vals_all)) + rmse_span * 0.20,
        ),
        "NSE": (
            max(0.0, float(np.min(nse_vals_all)) - nse_span * 0.12),
            min(1.0, float(np.max(nse_vals_all)) + nse_span * 0.24),
        ),
    }

    fig, axes = plt.subplots(2, 3, figsize=cfg["figsize"])
    fig.subplots_adjust(hspace=cfg["hspace"], wspace=cfg["wspace"])

    for r, (metric, lower_is_better) in enumerate(metrics):
        for c, site in enumerate(site_order):
            ax = axes[r, c]
            df = overall_dfs[site][["ModelCanon", "ModelDisplay", metric]].copy()
            df = df.dropna(subset=[metric]).sort_values(metric, ascending=lower_is_better).reset_index(drop=True)
            y = np.arange(len(df))
            vals = df[metric].to_numpy(dtype=float)
            colors = np.array([cfg["color_other"]] * len(df), dtype=object)

            ax.set_facecolor(cfg["panel_face"])
            for yi in range(len(df)):
                if yi % 2 == 1:
                    ax.axhspan(yi - 0.5, yi + 0.5, color=cfg["stripe_color"], alpha=0.55, zorder=0)

            main_mask = df["ModelCanon"] == PROPOSED_CANON
            base_mask = df["ModelCanon"] == site_best_baseline[site]
            colors[base_mask.values] = cfg["color_best"]
            colors[main_mask.values] = cfg["color_main"]

            ax.barh(
                y,
                vals,
                color=colors.tolist(),
                height=cfg["bar_height"],
                edgecolor="white",
                linewidth=0.8,
                alpha=0.95,
                zorder=2,
            )
            ax.scatter(vals, y, s=26, c=colors.tolist(), edgecolor="white", linewidth=0.8, zorder=3)
            ax.invert_yaxis()
            ax.grid(axis="x", linestyle=":", linewidth=1.0, color=cfg["grid_color"], alpha=0.35)
            ax.set_xlim(*row_xlim[metric])

            ax.set_yticks(y)
            rank_labels = [f"{i+1}. {name}" for i, name in enumerate(df["ModelDisplay"].tolist())]
            ax.set_yticklabels(rank_labels, fontsize=cfg["ytick_size"])
            ax.tick_params(axis="x", labelsize=cfg["xtick_size"])
            ax.tick_params(axis="y", length=0)
            if r == 0:
                ax.set_xlabel("RMSE", fontsize=cfg["label_size"], fontweight="bold")
            if r == 1:
                ax.set_xlabel("NSE", fontsize=cfg["label_size"], fontweight="bold")

            panel_idx = r * 3 + c
            ax.text(
                0.01,
                1.05,
                f"{panel_tags[panel_idx]} {site}",
                transform=ax.transAxes,
                ha="left",
                va="top",
                fontsize=cfg["panel_size"],
                fontweight="bold",
            )

            # 只对主模型与该站点最优基线做强调标注，其余模型由条形长度表达。
            for idx in np.where(main_mask.values | base_mask.values)[0]:
                x_val = vals[idx]
                ax.text(
                    x_val + (row_xlim[metric][1] - row_xlim[metric][0]) * 0.008,
                    idx,
                    f"{x_val:.3f}",
                    ha="left",
                    va="center",
                    fontsize=cfg["annot_size"],
                    fontweight="bold",
                    color="#111827",
                    bbox={"boxstyle": "round,pad=0.18", "facecolor": "white", "edgecolor": cfg["value_box_edge"], "alpha": 0.95},
                )

            ax.spines["top"].set_visible(False)
            ax.spines["right"].set_visible(False)
            ax.spines["left"].set_color(cfg["spine_color"])
            ax.spines["bottom"].set_color(cfg["spine_color"])

    legend_handles = [
        Patch(facecolor=cfg["color_main"], edgecolor="none", label="Ours"),
        Patch(facecolor=cfg["color_best"], edgecolor="none", label="Best baseline (site-specific)"),
        Patch(facecolor=cfg["color_other"], edgecolor="none", label="Other baselines"),
    ]
    fig.legend(
        handles=legend_handles,
        loc="upper center",
        ncol=3,
        frameon=False,
        bbox_to_anchor=(0.5, 0.945),
        fontsize=11,
        fancybox=True,
        framealpha=0.92,
        edgecolor=cfg["spine_color"],
        facecolor="#F8FAFC",
    )

    out = out_path or (FIG_DIR / "Overall_cross_site.tiff")
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=300)
    plt.close(fig)
    return out


# 手动执行（Figure 4）
# _set_global_style()
# FIG_DIR.mkdir(parents=True, exist_ok=True)
# overall_dfs, site_best_baseline, scenario_df, case_samples, ablation_delta_df = _prepare_data()
# out = plot_figure4_overall_cross_site(overall_dfs, site_best_baseline)
# print(f"[saved] {out}")


# =========================
# Figure 5
# =========================
def plot_figure5_scenario_overall(
    scenario_df: pd.DataFrame,
    out_path: Path | None = None,
) -> Path:
    """
    Figure 5:
    (a) Intensity scenarios heatmap + marginals
    (b) Duration scenarios heatmap + marginals
    """
    # ---- Figure 5 config (可单独调参) ----
    cfg = {
        "figsize": (16.2, 7.2),
        "width_ratios": [1.0, 1.0, 0.05],
        "wspace": 0.16,
        "cbar_shift_left": 0.015,
        "annot_size": 14,
        "annot_n_size": 13,
        "tick_size": 12,
        "panel_size": 13,
        "cbar_label_size": 12.5,
        "cbar_tick_size": 11.5,
        "cmap": "RdBu_r",
        "panel_face": "#f8fafc",
        "grid_color": "#d1d9e6",
        "title_box_face": "#e2e8f0",
        "na_face": "#f1f5f9",
        "na_edge": "#94a3b8",
    }

    int_df = scenario_df[scenario_df["ScenarioType"] == "Intensity"].pivot(
        index="Site",
        columns="Scenario",
        values="Delta_RMSE_pct",
    )
    int_n_df = scenario_df[scenario_df["ScenarioType"] == "Intensity"].pivot(
        index="Site",
        columns="Scenario",
        values="N_events",
    )
    dur_df = scenario_df[scenario_df["ScenarioType"] == "Duration"].pivot(
        index="Site",
        columns="Scenario",
        values="Delta_RMSE_pct",
    )
    dur_n_df = scenario_df[scenario_df["ScenarioType"] == "Duration"].pivot(
        index="Site",
        columns="Scenario",
        values="N_events",
    )

    int_df = int_df.reindex(index=["W2127", "W2128", "W2129"], columns=["Weak", "Moderate", "Strong"])
    int_n_df = int_n_df.reindex(index=["W2127", "W2128", "W2129"], columns=["Weak", "Moderate", "Strong"])
    dur_df = dur_df.reindex(index=["W2127", "W2128", "W2129"], columns=["Short", "Medium", "Long"])
    dur_n_df = dur_n_df.reindex(index=["W2127", "W2128", "W2129"], columns=["Short", "Medium", "Long"])

    values = np.concatenate([int_df.values.flatten(), dur_df.values.flatten()])
    values = values[~np.isnan(values)]
    lim = max(abs(values.min()), abs(values.max())) if len(values) > 0 else 1.0
    vmin, vmax = -lim, lim
    norm = TwoSlopeNorm(vmin=vmin, vcenter=0.0, vmax=vmax)

    fig = plt.figure(figsize=cfg["figsize"])
    gs = GridSpec(1, 3, figure=fig, width_ratios=cfg["width_ratios"], wspace=cfg["wspace"])

    left = gs[0, 0].subgridspec(2, 2, width_ratios=[1.0, 0.34], height_ratios=[1.0, 0.34], wspace=0.01, hspace=0.12)
    right = gs[0, 1].subgridspec(2, 2, width_ratios=[1.0, 0.34], height_ratios=[1.0, 0.34], wspace=0.01, hspace=0.12)

    ax0 = fig.add_subplot(left[0, 0])
    ax0_site = fig.add_subplot(left[0, 1], sharey=ax0)
    ax0_scen = fig.add_subplot(left[1, 0], sharex=ax0)
    ax0_empty = fig.add_subplot(left[1, 1])

    ax1 = fig.add_subplot(right[0, 0])
    ax1_site = fig.add_subplot(right[0, 1], sharey=ax1)
    ax1_scen = fig.add_subplot(right[1, 0], sharex=ax1)
    ax1_empty = fig.add_subplot(right[1, 1])

    cax = fig.add_subplot(gs[0, 2])
    cax_pos = cax.get_position()
    cax.set_position([cax_pos.x0 - cfg["cbar_shift_left"], cax_pos.y0, cax_pos.width, cax_pos.height])

    im0 = ax0.imshow(int_df.values.astype(float), cmap=cfg["cmap"], norm=norm, aspect="equal")
    im1 = ax1.imshow(dur_df.values.astype(float), cmap=cfg["cmap"], norm=norm, aspect="equal")

    panel_specs = [
        (ax0, ax0_site, ax0_scen, ax0_empty, int_df, int_n_df, "(a)", "By Precipitation Intensity"),
        (ax1, ax1_site, ax1_scen, ax1_empty, dur_df, dur_n_df, "(b)", "By Precipitation Duration"),
    ]

    for ax, ax_site, ax_scen, ax_empty, df, n_df, panel, ptitle in panel_specs:
        ax.set_facecolor(cfg["panel_face"])
        ax.set_xticks(np.arange(df.shape[1]))
        ax.set_xticklabels(df.columns, fontsize=cfg["tick_size"], fontweight="bold")
        ax.set_yticks(np.arange(df.shape[0]))
        ax.set_yticklabels(df.index, fontsize=cfg["tick_size"], fontweight="bold")
        ax.set_xticks(np.arange(-0.5, df.shape[1], 1), minor=True)
        ax.set_yticks(np.arange(-0.5, df.shape[0], 1), minor=True)
        ax.grid(which="minor", color=cfg["grid_color"], linestyle="-", linewidth=1.15)
        ax.tick_params(which="minor", bottom=False, left=False)
        ax.set_title(
            f"{panel} {ptitle}",
            loc="left",
            fontsize=cfg["panel_size"],
            fontweight="bold",
            pad=8,
        )
        arr = df.values.astype(float)
        n_arr = n_df.values.astype(float)

        if np.isfinite(arr).any():
            max_i, max_j = np.unravel_index(np.nanargmax(arr), arr.shape)
            min_i, min_j = np.unravel_index(np.nanargmin(arr), arr.shape)
        else:
            max_i = max_j = min_i = min_j = -1

        for i in range(arr.shape[0]):
            for j in range(arr.shape[1]):
                if np.isnan(arr[i, j]):
                    ax.add_patch(
                        Rectangle(
                            (j - 0.5, i - 0.5),
                            1.0,
                            1.0,
                            facecolor=cfg["na_face"],
                            edgecolor=cfg["na_edge"],
                            hatch="///",
                            linewidth=0.6,
                            zorder=3,
                        )
                    )
                    ax.text(j, i, "NA", ha="center", va="center", fontsize=cfg["annot_size"], color="#64748b", fontweight="bold")
                else:
                    val = arr[i, j]
                    txt_color = "white" if abs(val) > 0.58 * lim else "#111827"
                    n_evt = "NA" if np.isnan(n_arr[i, j]) else f"{int(round(n_arr[i, j]))}"
                    ax.text(
                        j,
                        i - 0.10,
                        f"{val:.1f}%",
                        ha="center",
                        va="center",
                        fontsize=cfg["annot_size"],
                        color=txt_color,
                        fontweight="bold",
                    )
                    ax.text(
                        j,
                        i + 0.20,
                        f"n={n_evt}",
                        ha="center",
                        va="center",
                        fontsize=cfg["annot_n_size"],
                        color=txt_color,
                    )

                if i == max_i and j == max_j:
                    ax.add_patch(Rectangle((j - 0.5, i - 0.5), 1.0, 1.0, fill=False, edgecolor="#0f172a", linewidth=1.8))
                if i == min_i and j == min_j:
                    ax.add_patch(
                        Rectangle((j - 0.5, i - 0.5), 1.0, 1.0, fill=False, edgecolor="#7f1d1d", linewidth=1.6, linestyle="--")
                    )

        for spine in ax.spines.values():
            spine.set_color("#cbd5e1")
            spine.set_linewidth(1.0)

        # 右侧：站点均值（行均值）
        site_mean = np.nanmean(arr, axis=1)
        ypos = np.arange(len(site_mean))
        site_colors = ["#2f6fae" if v >= 0 else "#c66a1a" for v in site_mean]
        ax_site.barh(ypos, site_mean, color=site_colors, height=0.52, alpha=0.9)
        ax_site.axvline(0.0, color="#94a3b8", lw=1.0)
        ax_site.set_ylim(ax.get_ylim())
        ax_site.set_yticks(ypos)
        ax_site.tick_params(axis="y", left=False, labelleft=False)
        ax_site.tick_params(axis="x", labelsize=9)
        ax_site.set_title("Site mean", fontsize=12, pad=6, fontweight="bold")
        ax_site.grid(axis="x", linestyle=":", color="#cbd5e1", alpha=0.6)
        for lbl in ax_site.get_xticklabels():
            lbl.set_fontweight("bold")
        for spine in ax_site.spines.values():
            spine.set_color("#cbd5e1")
            spine.set_linewidth(1.0)

        # 下方：场景均值（列均值）
        scen_mean = np.nanmean(arr, axis=0)
        xpos = np.arange(len(scen_mean))
        scen_colors = ["#2f6fae" if v >= 0 else "#c66a1a" for v in scen_mean]
        ax_scen.bar(xpos, scen_mean, color=scen_colors, width=0.50, alpha=0.9)
        ax_scen.axhline(0.0, color="#94a3b8", lw=1.0)
        ax_scen.set_xticks(xpos)
        ax_scen.set_xticklabels(df.columns, fontsize=10)
        ax_scen.tick_params(axis="y", labelsize=9)
        ax_scen.set_ylabel("Scenario mean", fontsize=12, fontweight="bold")
        ax_scen.grid(axis="y", linestyle=":", color="#cbd5e1", alpha=0.6)
        for lbl in ax_scen.get_xticklabels():
            lbl.set_fontweight("bold")
        for lbl in ax_scen.get_yticklabels():
            lbl.set_fontweight("bold")
        for spine in ax_scen.spines.values():
            spine.set_color("#cbd5e1")
            spine.set_linewidth(1.0)

        # 右下角空白区放说明，避免版面浪费
        ax_empty.axis("off")
        ax_empty.text(
            0.10,
            0.80,
            "Cell text:\nΔRMSE%\n& n events",
            transform=ax_empty.transAxes,
            ha="left",
            va="top",
            fontsize=10,
            fontweight="bold",
            color="#475569",
        )
        ax_empty.text(
            0.10,
            0.38,
            "Solid box: best\nDashed box: weakest",
            transform=ax_empty.transAxes,
            ha="left",
            va="top",
            fontsize=10,
            fontweight="bold",
            color="#475569",
        )

    cbar = fig.colorbar(im1, cax=cax)
    cbar.set_label("ΔRMSE (%) vs best baseline", fontsize=cfg["cbar_label_size"], fontweight="bold")
    cbar.ax.tick_params(labelsize=cfg["cbar_tick_size"])
    cbar.outline.set_edgecolor("#cbd5e1")
    cbar.outline.set_linewidth(1.0)

    out = out_path or (FIG_DIR / "Scenario_overall.tiff")
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=300)
    plt.close(fig)
    return out


# 手动执行（Figure 5）
# _set_global_style()
# FIG_DIR.mkdir(parents=True, exist_ok=True)
# overall_dfs, site_best_baseline, scenario_df, case_samples, ablation_delta_df = _prepare_data()
# out = plot_figure5_scenario_overall(scenario_df)
# print(f"[saved] {out}")


# =========================
# Figure 6
# =========================
def _phase_bounds(y_true: np.ndarray) -> Tuple[int, int]:
    n = len(y_true)
    if n <= 3:
        return max(1, n // 3), max(2, 2 * n // 3)
    peak_idx = int(np.argmax(y_true))
    peak_win = max(1, int(0.08 * n))
    onset_end = max(1, peak_idx - peak_win)
    peak_end = min(n - 1, peak_idx + peak_win)
    if onset_end >= peak_end:
        onset_end = max(1, n // 3)
        peak_end = max(onset_end + 1, 2 * n // 3)
    return onset_end, peak_end


def plot_figure6_event_mechanism(
    case_samples: List[Dict[str, object]] | None = None,
    out_path: Path | None = None,
    case_idx_w2127: int | None = None,
    case_idx_w2128: int | None = None,
    case_idx_w2129: int | None = None,
) -> Path:
    """
    Figure 6:
    3 个事件子图，右上角显示 RMSE/MAE 对比，图例放在子图 a 上方。
    可通过 case_idx_w2127/case_idx_w2128/case_idx_w2129 指定各站点个例编号（1-based）。
    """
    # ---- Figure 6 config (可单独调参) ----
    cfg = {
        "figsize": (12.0, 7.8),
        "hspace": 0.16,
        "line_main": 2.0,
        "line_base": 1.8,
        "line_truth": 2.0,
        "label_size": 15,
        "tick_size": 13,
        "legend_size": 13,
        "panel_size": 11,
        "metric_text_size": 10,
        "span_alpha": 0.45,
    }

    custom_case_map: Dict[str, int] = {}
    if case_idx_w2127 is not None:
        custom_case_map["W2127"] = int(case_idx_w2127)
    if case_idx_w2128 is not None:
        custom_case_map["W2128"] = int(case_idx_w2128)
    if case_idx_w2129 is not None:
        custom_case_map["W2129"] = int(case_idx_w2129)

    if case_samples is None or len(custom_case_map) > 0:
        site_best_baseline: Dict[str, str] = {}
        for site_full in SITE_FULL:
            site = SITE_CODE[site_full]
            odf = read_overall_site_table(site_full)
            site_best_baseline[site] = select_best_baseline(odf)
        case_samples = build_case_samples(site_best_baseline, case_index_by_site=custom_case_map or None)

    fig = plt.figure(figsize=cfg["figsize"])
    gs = GridSpec(3, 1, figure=fig, hspace=cfg["hspace"])

    for i, case in enumerate(case_samples):
        ax = fig.add_subplot(gs[i, 0])
        y_true = case["y_true"]
        y_main = case["y_main"]
        y_base = case["y_base"]
        x = np.arange(len(y_true))
        onset_end, peak_end = _phase_bounds(y_true)

        ax.axvspan(0, onset_end, color="#dbeafe", alpha=cfg["span_alpha"], lw=0)
        ax.axvspan(onset_end, peak_end, color="#fde68a", alpha=cfg["span_alpha"], lw=0)
        ax.axvspan(peak_end, len(y_true), color="#e5e7eb", alpha=cfg["span_alpha"], lw=0)

        ax.plot(x, y_true, color="black", lw=cfg["line_truth"], label="Observation")
        ax.plot(x, y_main, color="#2f6fae", lw=cfg["line_main"], label="Ours")
        ax.plot(
            x,
            y_base,
            color="#d97706",
            lw=cfg["line_base"],
            ls="--",
            label="Best baseline (site-specific)",
        )

        ax.text(0.01,0.95,f"({chr(97+i)}) {case['case_id']}",
            transform=ax.transAxes,
            ha="left",
            va="top",
            fontsize=cfg["panel_size"],
            fontweight="bold",
        )
        ax.text(0.01,0.83,f"Best baseline: {case['baseline_display']}",
            transform=ax.transAxes,
            ha="left",
            va="top",
            fontsize=cfg["panel_size"] - 1,
            color="#d97706",
        )
        ax.text(0.99,0.95,
            (f"RMSE {case['rmse_main']:.3f} vs {case['rmse_base']:.3f} (Δ{case['delta_rmse_pct']:.1f}%)\n"),
            transform=ax.transAxes,
            ha="right",
            va="top",
            fontsize=cfg["metric_text_size"],
        )
        ax.text(0.99,0.85,
            (f"MAE  {case['mae_main']:.3f} vs {case['mae_base']:.3f} (Δ{case['delta_mae_pct']:.1f}%)"),
            transform=ax.transAxes,
            ha="right",
            va="top",
            fontsize=cfg["metric_text_size"],
        )


        ax.set_ylabel("Rain rate (mm/h)", fontsize=cfg["label_size"], fontweight="bold")
        ax.tick_params(axis="both", labelsize=cfg["tick_size"])
        if i == len(case_samples) - 1:
            ax.set_xlabel("Timeline index (min)", fontsize=cfg["label_size"], fontweight="bold")
        if i == 0:
            ax.legend(
                frameon=False,
                ncol=3,
                loc="lower center",
                bbox_to_anchor=(0.5, 1.01),
                borderaxespad=0.0,
                fontsize=cfg["legend_size"],
            )

    out = out_path or (FIG_DIR / "Event_mechanism.tiff")
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=300)
    plt.close(fig)
    return out


# 手动执行（Figure 6）
# _set_global_style()
# FIG_DIR.mkdir(parents=True, exist_ok=True)
# overall_dfs, site_best_baseline, scenario_df, case_samples, ablation_delta_df = _prepare_data()
# out = plot_figure6_event_mechanism(case_samples)
# out = plot_figure6_event_mechanism(case_idx_w2127=8, case_idx_w2128=8, case_idx_w2129=4)
# print(f"[saved] {out}")


# =========================
# Figure 7
# =========================
def plot_figure7_ablation(
    ablation_delta_df: pd.DataFrame,
    out_path: Path | None = None,
) -> Path:
    """
    Figure 7:
    2x2 compact matrix readouts:
    (a) RMSE, (b) MAE, (c) MSE, (d) NSE
    """
    # ---- Figure 7 config (可单独调参) ----
    cfg = {
        "figsize": (14.4, 10.2),
        "wspace": 0.20,
        "hspace": 0.17,
        "annot_size": 11.5,
        "tick_size": 13,
        "panel_size": 13,
        "matrix_cmap": "RdBu_r",
        "title_box_face": "#e2e8f0",
        "grid_color": "#cbd5e1",
    }

    df = ablation_delta_df.copy()
    df["ModelDisplayPlot"] = np.where(df["ModelCanon"] == PROPOSED_CANON, "Full", df["ModelDisplay"])

    # Use RMSE degradation ranking to keep a consistent order across all 4 metric panels.
    rmse_rank = (
        df.groupby("ModelDisplayPlot")["Delta_RMSE_vs_Proposed_pct"]
        .mean()
        .sort_values(ascending=False)
    )
    model_order = [m for m in rmse_rank.index.tolist() if m != "Full"]
    if "Full" in rmse_rank.index:
        model_order = ["Full"] + model_order

    model_ids = {"Full": "A0"}
    next_idx = 1
    for name in model_order:
        if name == "Full":
            continue
        model_ids[name] = f"A{next_idx}"
        next_idx += 1

    metric_specs = [
        ("Delta_RMSE_vs_Proposed_pct", "RMSE", "(a)", "ΔRMSE (%) vs Full", "{:+.1f}%"),
        ("Delta_MAE_vs_Proposed_pct", "MAE", "(b)", "ΔMAE (%) vs Full", "{:+.1f}%"),
        ("Delta_MSE_vs_Proposed_pct", "MSE", "(c)", "ΔMSE (%) vs Full", "{:+.1f}%"),
        ("Delta_NSE_vs_Proposed_pp", "NSE", "(d)", "ΔNSE (pp, loss) vs Full", "{:+.2f}"),
    ]

    fig, axes = plt.subplots(2, 2, figsize=cfg["figsize"])
    fig.subplots_adjust(wspace=cfg["wspace"], hspace=cfg["hspace"])
    for ax, (col, metric_name, panel_tag, cbar_label, fmt) in zip(axes.flat, metric_specs):
        pivot = df.pivot(index="ModelDisplayPlot", columns="Site", values=col)
        pivot = pivot.reindex(index=model_order, columns=["W2127", "W2128", "W2129"])
        pivot_id = pivot.copy()
        pivot_id.index = [model_ids[m] for m in pivot.index]

        arr = pivot_id.values.astype(float)
        vals = arr.flatten()
        vals = vals[~np.isnan(vals)]
        lim = max(abs(vals.min()), abs(vals.max())) if len(vals) > 0 else 1.0
        norm = TwoSlopeNorm(vmin=-lim, vcenter=0.0, vmax=lim)

        im = ax.imshow(arr, cmap=cfg["matrix_cmap"], norm=norm, aspect="auto")
        ax.set_xticks(np.arange(pivot_id.shape[1]))
        ax.set_xticklabels(pivot_id.columns, fontsize=cfg["tick_size"], fontweight="bold")
        ax.set_yticks(np.arange(pivot_id.shape[0]))
        ax.set_yticklabels(pivot_id.index, fontsize=cfg["tick_size"], fontweight="bold")
        ax.set_xticks(np.arange(-0.5, pivot_id.shape[1], 1), minor=True)
        ax.set_yticks(np.arange(-0.5, pivot_id.shape[0], 1), minor=True)
        ax.grid(which="minor", color=cfg["grid_color"], linestyle="-", linewidth=1.0)
        ax.tick_params(which="minor", bottom=False, left=False)
        ax.set_title(
            f"{panel_tag} {metric_name}",
            loc="left",
            fontsize=cfg["panel_size"],
            fontweight="bold",
            pad=8,
        )
        for i in range(arr.shape[0]):
            for j in range(arr.shape[1]):
                if np.isnan(arr[i, j]):
                    txt, color = "NA", "#64748b"
                else:
                    txt = fmt.format(arr[i, j])
                    color = "white" if abs(arr[i, j]) > 0.58 * lim else "#111827"
                ax.text(j, i, txt, ha="center", va="center", fontsize=cfg["annot_size"], color=color, fontweight="bold")
        cbar = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.03)
        cbar.set_label(cbar_label, fontsize=12.0, fontweight="bold")
        cbar.ax.tick_params(labelsize=10)

    mapping_lines = [f"{model_ids[name]}: {name}" for name in model_order]
    fig.text(
        0.5,
        0.056,
        " | ".join(mapping_lines),
        ha="center",
        va="bottom",
        fontsize=11,
        fontweight="bold",
        color="#334155",
    )

    out = out_path or (FIG_DIR / "Figure7_Ablation_contribution.tiff")
    out.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(out, dpi=300)
    plt.close(fig)
    return out


# 手动执行（Figure 7）
# _set_global_style()
# FIG_DIR.mkdir(parents=True, exist_ok=True)
# overall_dfs, site_best_baseline, scenario_df, case_samples, ablation_delta_df = _prepare_data()
# out = plot_figure7_ablation(ablation_delta_df)
# print(f"[saved] {out}")
