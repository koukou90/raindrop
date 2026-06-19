"""Update metrics_comparison_template xlsx from latest experiment outputs."""

from __future__ import annotations

import argparse
import re
from pathlib import Path

from openpyxl import load_workbook


OVERALL_METRICS = ['RMSE', 'MAE', 'MSE', 'CC', 'R2', 'MAPE', 'Bias', 'NSE', 'Peak-MAE']
STEP_PREFIXES = ['T+1', 'T+2', 'T+3', 'T+4', 'T+5']

MODEL_ALIASES = {
    'TripleStreamV3': 'TripleStreamV3_self_distill',
    'BaselineLinear': 'BaselineDLinear',
}

NEW_BASELINES = [
    'BaselineDLinear',
    'BaselinePatchTST',
    'BaselineTiDE',
    'BaselineCNNTransformer',
    'BaselineiTransformer',
]

DEFAULT_MODEL_ORDER = [
    'TripleStreamV3_self_distill',
    'BaselineLightGBM',
    'BaselineMLP',
    'BaselineCNNTransformer',
    'BaselineLSTM',
    'BaselineiTransformer',
    'BaselinePatchTST',
    'BaselineDLinear',
    'BaselineTiDE',
]


def _parse_metric_row(values: list[str]) -> dict[str, float]:
    if len(values) < len(OVERALL_METRICS):
        raise ValueError(f'invalid metric row: {values!r}')
    return {key: float(val) for key, val in zip(OVERALL_METRICS, values[: len(OVERALL_METRICS)])}


def parse_test_metrics(path: Path) -> dict:
    text = path.read_text(encoding='utf-8')
    result = {'overall': {}, 'steps': {}}

    overall_block = re.search(
        r'整体评价指标（时间点级，论文主口径）:\s*-+\s*'
        r'RMSE\s+MAE\s+MSE\s+CC\s+R[²2]\s+MAPE\s+Bias\s+NSE\s+Peak-MAE\s*-+\s*'
        r'([^\n]+)',
        text,
    )
    if overall_block:
        result['overall'] = _parse_metric_row(overall_block.group(1).split())

    table_block = re.search(
        r'时间点级逐步评价指标（表格形式，便于对比不同步长）:\s*-+\s*'
        r'Step\s+RMSE\s+MAE\s+MSE\s+CC\s+R[²2]\s+MAPE\s+Bias\s+NSE\s+Peak-MAE\s*-+\s*'
        r'(.*?)(?:\n\s*\n|\Z)',
        text,
        re.S,
    )
    if table_block:
        for line in table_block.group(1).splitlines():
            line = line.strip()
            if not line or not line.startswith('T+'):
                continue
            step_name, *values = line.split()
            if step_name in STEP_PREFIXES:
                result['steps'][step_name] = _parse_metric_row(values)

    return result


def find_latest_run(site_dir: Path, model_name: str) -> Path | None:
    for folder in sorted(site_dir.glob(f'{model_name}_*'), key=lambda p: p.stat().st_mtime, reverse=True):
        metrics_file = folder / 'test_metrics.txt'
        if metrics_file.exists():
            return metrics_file
    return None


def collect_site_metrics(site_dir: Path) -> dict[str, dict]:
    source_models = [
        'TripleStreamV3',
        'BaselineLightGBM',
        'BaselineMLP',
        'BaselineLSTM',
        *NEW_BASELINES,
    ]
    metrics_by_model: dict[str, dict] = {}
    for model_name in source_models:
        metrics_file = find_latest_run(site_dir, model_name)
        if metrics_file is None:
            print(f'[skip] missing metrics for {model_name}')
            continue
        display_name = MODEL_ALIASES.get(model_name, model_name)
        metrics_by_model[display_name] = parse_test_metrics(metrics_file)
        print(f'[ok] {display_name} <= {metrics_file.parent.name}')
    return metrics_by_model


def rebuild_overall_sheet(ws, models: list[str], metrics_by_model: dict[str, dict]):
    ws.delete_rows(1, ws.max_row)
    row = 1

    ws.cell(row, 1, 'Overall')
    row += 1
    ws.cell(row, 1, 'Model')
    for col, metric in enumerate(OVERALL_METRICS, start=2):
        ws.cell(row, col, metric)
    row += 1
    for model in models:
        ws.cell(row, 1, model)
        overall = metrics_by_model.get(model, {}).get('overall', {})
        for col, metric in enumerate(OVERALL_METRICS, start=2):
            ws.cell(row, col, overall.get(metric))
        row += 1

    row += 1

    for metric_key in ['RMSE', 'MAE', 'MAPE']:
        ws.cell(row, 1, metric_key)
        row += 1
        ws.cell(row, 1, 'Model')
        for col, step in enumerate(STEP_PREFIXES, start=2):
            ws.cell(row, col, step)
        row += 1
        for model in models:
            ws.cell(row, 1, model)
            steps = metrics_by_model.get(model, {}).get('steps', {})
            for col, step in enumerate(STEP_PREFIXES, start=2):
                ws.cell(row, col, steps.get(step, {}).get(metric_key))
            row += 1
        row += 1


def update_xlsx(xlsx_path: Path, site_dir: Path, model_order: list[str]):
    metrics_by_model = collect_site_metrics(site_dir)
    models = [model for model in model_order if model in metrics_by_model]

    wb = load_workbook(xlsx_path)
    ws = wb['Overall']
    rebuild_overall_sheet(ws, models, metrics_by_model)
    wb.save(xlsx_path)
    print(f'[saved] {xlsx_path}')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--site_name', required=True)
    parser.add_argument('--out_path', default='out')
    args = parser.parse_args()

    site_dir = Path(args.out_path) / args.site_name
    xlsx_path = site_dir / f'metrics_comparison_template_{args.site_name}.xlsx'
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    update_xlsx(xlsx_path, site_dir, DEFAULT_MODEL_ORDER)


if __name__ == '__main__':
    main()
