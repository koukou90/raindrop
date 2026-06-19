"""Update ablation metrics xlsx (A-class only) from experiment outputs."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook

from tools.update_metrics_xlsx import OVERALL_METRICS, STEP_PREFIXES, parse_test_metrics


SITES = ['W2127_Haichaoba', 'W2128_Haichaoyinsi', 'W2129_buligou']
ABLATION_ROOT = '消融实验'

A_MODELS = [
    ('TripleStreamV3_self_distill', 'TripleStreamV3', None),
    ('AblateV3NoPersistence', 'AblateV3NoPersistence', 'A'),
    ('AblateV3NoMixGate', 'AblateV3NoMixGate', 'A'),
    ('AblateV3NoBinAware', 'AblateV3NoBinAware', 'A'),
    ('AblateV3NoAux', 'AblateV3NoAux', 'A'),
    ('AblateV3NoStreamGate', 'AblateV3NoStreamGate', 'A'),
]


def find_latest_metrics(site_dir: Path, source_name: str, ablation_kind: str | None) -> Path | None:
    if ablation_kind is None:
        search_dir = site_dir
    else:
        search_dir = site_dir / ABLATION_ROOT / ablation_kind

    if not search_dir.exists():
        return None

    for folder in sorted(search_dir.glob(f'{source_name}_*'), key=lambda p: p.stat().st_mtime, reverse=True):
        metrics_file = folder / 'test_metrics.txt'
        if metrics_file.exists():
            return metrics_file
    return None


def collect_ablation_metrics(site_dir: Path, model_defs: list[tuple]) -> dict[str, dict]:
    metrics_by_model: dict[str, dict] = {}
    for display_name, source_name, ablation_kind in model_defs:
        metrics_file = find_latest_metrics(site_dir, source_name, ablation_kind)
        if metrics_file is None:
            print(f'[skip] {display_name}')
            continue
        metrics_by_model[display_name] = parse_test_metrics(metrics_file)
        print(f'[ok] {display_name} <= {metrics_file.parent.name}')
    return metrics_by_model


def write_section(ws, start_row: int, title: str, models: list[str], metrics_by_model: dict[str, dict]) -> int:
    row = start_row
    ws.cell(row, 1, title)
    row += 1

    ws.cell(row, 1, 'Model')
    for col, metric in enumerate(OVERALL_METRICS, start=2):
        ws.cell(row, col, metric)
    row += 1
    for model in models:
        ws.cell(row, 1, model)
        overall = metrics_by_model.get(model, {}).get('overall', {})
        for col, metric in enumerate(OVERALL_METRICS, start=2):
            ws.cell(row, col, overall.get(metric))
        row += 1

    row += 1
    for metric_key in ['RMSE', 'MAE', 'MAPE']:
        ws.cell(row, 1, metric_key)
        row += 1
        ws.cell(row, 1, 'Model')
        for col, step in enumerate(STEP_PREFIXES, start=2):
            ws.cell(row, col, step)
        row += 1
        for model in models:
            ws.cell(row, 1, model)
            steps = metrics_by_model.get(model, {}).get('steps', {})
            for col, step in enumerate(STEP_PREFIXES, start=2):
                ws.cell(row, col, steps.get(step, {}).get(metric_key))
            row += 1
        row += 1
    return row


def update_ablation_xlsx(site_name: str, out_path: str = 'out'):
    site_dir = Path(out_path) / site_name
    xlsx_path = site_dir / ABLATION_ROOT / f'metrics_ablation_{site_name}.xlsx'

    a_metrics = collect_ablation_metrics(site_dir, A_MODELS)
    a_models = [name for name, *_ in A_MODELS if name in a_metrics]

    wb = Workbook()
    ws = wb.active
    ws.title = 'Ablation'
    write_section(ws, 1, 'A类结构消融', a_models, a_metrics)

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)
    print(f'[saved] {xlsx_path}')


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--site_name', choices=SITES + ['all'], default='all')
    parser.add_argument('--out_path', default='out')
    args = parser.parse_args()

    sites = SITES if args.site_name == 'all' else [args.site_name]
    for site in sites:
        update_ablation_xlsx(site, args.out_path)


if __name__ == '__main__':
    main()
