"""Generate section 4.2-4.6 tables only (Overall-only protocol)."""

from __future__ import annotations

import pickle
from pathlib import Path
from typing import Dict, List, Tuple
import sys

import numpy as np
import pandas as pd
from scipy import stats
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from utils.metrics import mae, mse, nse, peak_mae, rmse
from data_loader.dataset import prepare_datasets, split_by_events

OUT_DIR = ROOT / "out"
FIG_DIR = OUT_DIR / "figs"
DATA_DIR = ROOT / "data" / "图像preview3_Gap15_Len30"

SITES = ["W2127_Haichaoba", "W2128_Haichaoyinsi", "W2129_buligou"]
PROPOSED_MODEL = "TripleStreamV3_self_distill"
METRIC_COLS = ["RMSE", "MAE", "MSE", "NSE", "Peak-MAE"]
EXCLUDED_MODELS = {"BaselineTiDE", "TiDE"}

SITE_LABEL = {
    "W2127_Haichaoba": "W2127",
    "W2128_Haichaoyinsi": "W2128",
    "W2129_buligou": "W2129",
}

SITE_META = {
    "W2127_Haichaoba": {
        "StationName": "Haichaoba",
        "Region": "Minle, Qilian Mountains foothill",
        "ApproxElevation_m": 1500,
        "Terrain": "Mountain-front plain transition",
        "ClimateBackground": "Semi-arid continental monsoon edge",
        "PrecipitationType": "Convective-dominant warm-season rainfall",
    },
    "W2128_Haichaoyinsi": {
        "StationName": "Haichaoyinsi",
        "Region": "Minle, Qilian Mountains foothill",
        "ApproxElevation_m": 1500,
        "Terrain": "Hill-slope and mountain-front transition",
        "ClimateBackground": "Semi-arid continental monsoon edge",
        "PrecipitationType": "Convective and mixed stratiform rainfall",
    },
    "W2129_buligou": {
        "StationName": "Buligou",
        "Region": "Minle, Qilian Mountains foothill",
        "ApproxElevation_m": 1500,
        "Terrain": "Valley-gully terrain",
        "ClimateBackground": "Semi-arid continental monsoon edge",
        "PrecipitationType": "Short-duration convective rainfall",
    },
}

MODEL_LABEL = {
    "TripleStreamV3_self_distill": "Ours",
    "BaselineLightGBM": "LightGBM",
    "BaselineMLP": "MLP",
    "BaselineLSTM": "LSTM",
    "BaselineDLinear": "DLinear",
    "BaselineLinear": "DLinear",
    "BaselinePatchTST": "PatchTST",
    "BaselineTiDE": "TiDE",
    "BaselineCNNTransformer": "CNN-Transformer",
    "BaselineiTransformer": "iTransformer",
    "AblateV3NoPersistence": "NoPersistence",
    "AblateV3NoMixGate": "NoMixGate",
    "AblateV3NoBinAware": "NoBinAware",
    "AblateV3NoAux": "NoAux",
    "AblateV3NoStreamGate": "NoStreamGate",
}


def site_tag(site_name: str) -> str:
    return SITE_LABEL.get(site_name, site_name)


def model_tag(model_name: str) -> str:
    return MODEL_LABEL.get(model_name, model_name)


def _to_float(x):
    try:
        return float(x)
    except Exception:
        return np.nan


def _load_site_params(site: str) -> pd.DataFrame:
    params_path = DATA_DIR / site / f"{site}_params.csv"
    if not params_path.exists():
        raise FileNotFoundError(f"Missing params csv: {params_path}")
    df = pd.read_csv(params_path)
    if "Timestamp" in df.columns:
        df["Timestamp"] = pd.to_datetime(df["Timestamp"], errors="coerce")
    return df


def build_table1_site_info() -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    for site in SITES:
        meta = SITE_META[site]
        df = _load_site_params(site)
        rain_rate = pd.to_numeric(df["RainRate"], errors="coerce")
        train_events, val_events, test_events = split_by_events(
            pd.DataFrame({"RainRate": rain_rate}), train_ratio=0.8, val_ratio=0.1
        )
        rows.append(
            {
                "SiteCode": site_tag(site),
                "StationName": meta["StationName"],
                "Region": meta["Region"],
                "Coordinate": "TBD",
                "ApproxElevation_m": meta["ApproxElevation_m"],
                "Terrain": meta["Terrain"],
                "ClimateBackground": meta["ClimateBackground"],
                "PrecipitationType": meta["PrecipitationType"],
                "DataStart": df["Timestamp"].min(),
                "DataEnd": df["Timestamp"].max(),
                "TotalMinutes": int(len(df)),
                "ValidRainSeriesPoints": int(rain_rate.notna().sum()),
                "EventCount": int(len(train_events) + len(val_events) + len(test_events)),
            }
        )
    return pd.DataFrame(rows).sort_values("SiteCode").reset_index(drop=True)


def build_table2_variable_definition() -> pd.DataFrame:
    rows = [
        {
            "Category": "Raw DSD input",
            "Variable": "N(D_i), i=1..32",
            "Unit": "instrument native (number concentration spectrum)",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Drop number concentration in 32 diameter bins",
            "ModelUsage": "Input stream: conc",
        },
        {
            "Category": "Raw DSD input",
            "Variable": "V(D_i), i=1..32",
            "Unit": "m/s",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Drop fall velocity in 32 diameter bins",
            "ModelUsage": "Input stream: vel",
        },
        {
            "Category": "Derived physics",
            "Variable": "RainRate",
            "Unit": "mm/h",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Rainfall intensity derived from DSD",
            "ModelUsage": "Input phys[0], prediction target",
        },
        {
            "Category": "Derived physics",
            "Variable": "Dm",
            "Unit": "mm",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Mass-weighted mean drop diameter",
            "ModelUsage": "Input phys[1], auxiliary target",
        },
        {
            "Category": "Derived physics",
            "Variable": "LogNw",
            "Unit": "log10(mm^-1 m^-3)",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Logarithm of normalized intercept parameter",
            "ModelUsage": "Input phys[2], auxiliary target",
        },
        {
            "Category": "Derived physics",
            "Variable": "LWC",
            "Unit": "g/m^3",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Liquid water content",
            "ModelUsage": "Input phys[3], auxiliary target",
        },
        {
            "Category": "Derived physics",
            "Variable": "Z",
            "Unit": "mm^6/m^3",
            "TemporalResolution": "1 min",
            "Definition_or_Rule": "Radar reflectivity factor",
            "ModelUsage": "Input phys[4], auxiliary target",
        },
        {
            "Category": "Quality control",
            "Variable": "Event merge gap",
            "Unit": "min",
            "TemporalResolution": "N/A",
            "Definition_or_Rule": "Dry gaps <= 15 min are merged into one event",
            "ModelUsage": "Event identification rule",
        },
        {
            "Category": "Quality control",
            "Variable": "Minimum event duration",
            "Unit": "min",
            "TemporalResolution": "N/A",
            "Definition_or_Rule": "Events shorter than 30 min are removed",
            "ModelUsage": "Event filtering rule",
        },
        {
            "Category": "Sample construction",
            "Variable": "Sliding window",
            "Unit": "(seq_len, pred_len, stride)",
            "TemporalResolution": "min",
            "Definition_or_Rule": "(10, 5, 1), event-wise sampling without crossing boundaries",
            "ModelUsage": "Supervised sample generation",
        },
        {
            "Category": "Data split",
            "Variable": "Train/Val/Test",
            "Unit": "ratio",
            "TemporalResolution": "event-level",
            "Definition_or_Rule": "0.8 / 0.1 / 0.1 in chronological order by events",
            "ModelUsage": "Model training and evaluation protocol",
        },
    ]
    return pd.DataFrame(rows)


def _safe_len(ds) -> int:
    return 0 if ds is None else int(len(ds))


def build_table3_sample_statistics() -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    for site in SITES:
        df = _load_site_params(site)
        rain_rate = pd.to_numeric(df["RainRate"], errors="coerce")
        dm = pd.to_numeric(df["Dm"], errors="coerce")
        lognw = pd.to_numeric(df["LogNw"], errors="coerce")
        lwc = pd.to_numeric(df["LWC"], errors="coerce")
        z = pd.to_numeric(df["Z"], errors="coerce")

        datasets, _, events_info = prepare_datasets(
            site_name=site,
            data_dir=str(DATA_DIR),
            seq_len=10,
            pred_len=5,
            stride=1,
            train_ratio=0.8,
            val_ratio=0.1,
        )

        rr_valid = rain_rate.dropna()
        rows.append(
            {
                "Site": site_tag(site),
                "EventCount_train": len(events_info["train_events"]),
                "EventCount_val": len(events_info["val_events"]),
                "EventCount_test": len(events_info["test_events"]),
                "EventCount_total": len(events_info["train_events"]) + len(events_info["val_events"]) + len(events_info["test_events"]),
                "SampleCount_train": _safe_len(datasets.get("train")),
                "SampleCount_val": _safe_len(datasets.get("val")),
                "SampleCount_test": _safe_len(datasets.get("test")),
                "SampleCount_total": _safe_len(datasets.get("train")) + _safe_len(datasets.get("val")) + _safe_len(datasets.get("test")),
                "ValidRainSeriesPoints": int(rain_rate.notna().sum()),
                "RainyPoints_gt0": int((rain_rate.fillna(0.0) > 0).sum()),
                "RainRate_mean": float(rr_valid.mean()),
                "RainRate_median": float(rr_valid.median()),
                "RainRate_p95": float(rr_valid.quantile(0.95)),
                "RainRate_max": float(rr_valid.max()),
                "Dm_mean": float(dm.mean()),
                "LogNw_mean": float(lognw.mean()),
                "LWC_mean": float(lwc.mean()),
                "Z_mean": float(z.mean()),
            }
        )
    return pd.DataFrame(rows).sort_values("Site").reset_index(drop=True)


def read_overall_table(site: str) -> pd.DataFrame:
    xlsx = OUT_DIR / site / f"metrics_comparison_template_{site}.xlsx"
    wb = load_workbook(xlsx, data_only=True)
    ws = wb["Overall"]

    headers = [ws.cell(2, c).value for c in range(1, 7)]
    rows: List[List[object]] = []
    r = 3
    while True:
        model = ws.cell(r, 1).value
        if model is None:
            break
        row = [ws.cell(r, c).value for c in range(1, 7)]
        rows.append(row)
        r += 1

    df = pd.DataFrame(rows, columns=headers)
    df = df[~df["Model"].astype(str).str.strip().isin(EXCLUDED_MODELS)].reset_index(drop=True)
    df.insert(0, "Site", site)
    for col in METRIC_COLS:
        df[col] = df[col].map(_to_float)
    return df


def read_ablation_table(site: str) -> pd.DataFrame:
    pattern = f"metrics_ablation_{site}.xlsx"
    matches = list((OUT_DIR / site).rglob(pattern))
    if not matches:
        raise FileNotFoundError(f"Cannot find {pattern}")

    wb = load_workbook(matches[0], data_only=True)
    ws = wb["Ablation"]
    headers = [ws.cell(2, c).value for c in range(1, 7)]
    rows: List[List[object]] = []
    r = 3
    while True:
        model = ws.cell(r, 1).value
        if model is None:
            break
        rows.append([ws.cell(r, c).value for c in range(1, 7)])
        r += 1

    df = pd.DataFrame(rows, columns=headers)
    df.insert(0, "Site", site)
    for col in METRIC_COLS:
        df[col] = df[col].map(_to_float)
    return df


def select_best_baseline(df: pd.DataFrame) -> str:
    baseline_df = df[df["Model"] != PROPOSED_MODEL].copy()
    return baseline_df.sort_values("RMSE").iloc[0]["Model"]


def _model_prefixes(model_name: str) -> List[str]:
    if model_name.startswith("TripleStreamV3"):
        return ["TripleStreamV3"]
    if model_name == "BaselineDLinear":
        return ["BaselineDLinear", "BaselineLinear"]
    return [model_name]


def find_run_dir(site: str, model_name: str) -> Path:
    site_dir = OUT_DIR / site
    candidates: List[Path] = []
    for prefix in _model_prefixes(model_name):
        candidates.extend([p for p in site_dir.iterdir() if p.is_dir() and p.name.startswith(prefix)])
    candidates.sort(key=lambda p: p.stat().st_mtime, reverse=True)

    for folder in candidates:
        pred_pkl = folder / "timeline_predictions_by_event.pkl"
        label_pkl = folder / "timeline_labels_by_event.pkl"
        if pred_pkl.exists() and label_pkl.exists():
            return folder
    raise FileNotFoundError(f"No event-level output found for {site} / {model_name}")


def load_event_series(run_dir: Path) -> Tuple[List[np.ndarray], List[np.ndarray]]:
    pred = pickle.load(open(run_dir / "timeline_predictions_by_event.pkl", "rb"))
    label = pickle.load(open(run_dir / "timeline_labels_by_event.pkl", "rb"))

    if isinstance(pred, dict):
        keys = list(pred.keys())
        pred_list = [np.asarray(pred[k]).flatten() for k in keys]
        label_list = [np.asarray(label[k]).flatten() for k in keys]
        return pred_list, label_list

    pred_list = [np.asarray(x).flatten() for x in pred]
    label_list = [np.asarray(x).flatten() for x in label]
    return pred_list, label_list


def compute_overall_metrics(y_true: np.ndarray, y_pred: np.ndarray) -> Dict[str, float]:
    return {
        "RMSE": rmse(y_true, y_pred),
        "MAE": mae(y_true, y_pred),
        "MSE": mse(y_true, y_pred),
        "NSE": nse(y_true, y_pred),
        "Peak-MAE": peak_mae(y_true, y_pred, threshold=10.0),
    }


def _intensity_group(peak: float) -> str:
    if peak < 2.0:
        return "Weak (<2 mm/h)"
    if peak < 10.0:
        return "Moderate (2-10 mm/h)"
    return "Strong (>=10 mm/h)"


def _duration_group(length: int) -> str:
    if length < 60:
        return "Short (<60 min)"
    if length < 180:
        return "Medium (60-180 min)"
    return "Long (>=180 min)"


def build_scenario_table(site_best_baseline: Dict[str, str]) -> pd.DataFrame:
    records: List[Dict[str, object]] = []
    scenario_containers: Dict[Tuple[str, str, str], Dict[str, List[np.ndarray]]] = {}

    for site in SITES:
        baseline_name = site_best_baseline[site]
        main_run = find_run_dir(site, PROPOSED_MODEL)
        base_run = find_run_dir(site, baseline_name)
        pred_main, label_main = load_event_series(main_run)
        pred_base, label_base = load_event_series(base_run)
        n_events = min(len(pred_main), len(pred_base), len(label_main), len(label_base))

        for i in range(n_events):
            y_true = np.asarray(label_main[i]).flatten()
            y_main = np.asarray(pred_main[i]).flatten()
            y_base = np.asarray(pred_base[i]).flatten()
            if len(y_true) == 0:
                continue

            intensity = _intensity_group(float(np.max(y_true)))
            duration = _duration_group(int(len(y_true)))

            for s_type, s_group in [("Intensity", intensity), ("Duration", duration)]:
                key = (site, s_type, s_group)
                if key not in scenario_containers:
                    scenario_containers[key] = {
                        "true": [],
                        "main": [],
                        "base": [],
                    }
                scenario_containers[key]["true"].append(y_true)
                scenario_containers[key]["main"].append(y_main)
                scenario_containers[key]["base"].append(y_base)

    for (site, s_type, s_group), pack in scenario_containers.items():
        y_true = np.concatenate(pack["true"])
        y_main = np.concatenate(pack["main"])
        y_base = np.concatenate(pack["base"])
        m_main = compute_overall_metrics(y_true, y_main)
        m_base = compute_overall_metrics(y_true, y_base)

        rec: Dict[str, object] = {
            "Site": site_tag(site),
            "ScenarioType": s_type,
            "Scenario": s_group,
            "N_events": len(pack["true"]),
            "N_points": len(y_true),
            "BestBaseline": model_tag(site_best_baseline[site]),
        }
        for col in METRIC_COLS:
            rec[f"{col}_Proposed"] = m_main[col]
            rec[f"{col}_Baseline"] = m_base[col]

        rec["Delta_RMSE_pct"] = (m_base["RMSE"] - m_main["RMSE"]) / m_base["RMSE"] * 100.0
        rec["Delta_MAE_pct"] = (m_base["MAE"] - m_main["MAE"]) / m_base["MAE"] * 100.0
        rec["Delta_MSE_pct"] = (m_base["MSE"] - m_main["MSE"]) / m_base["MSE"] * 100.0
        rec["Delta_NSE"] = m_main["NSE"] - m_base["NSE"]
        if np.isnan(m_main["Peak-MAE"]) or np.isnan(m_base["Peak-MAE"]) or m_base["Peak-MAE"] == 0:
            rec["Delta_PeakMAE_pct"] = np.nan
        else:
            rec["Delta_PeakMAE_pct"] = (m_base["Peak-MAE"] - m_main["Peak-MAE"]) / m_base["Peak-MAE"] * 100.0
        records.append(rec)

    df = pd.DataFrame(records).sort_values(["Site", "ScenarioType", "Scenario"]).reset_index(drop=True)
    return df


def build_case_table(site_best_baseline: Dict[str, str]) -> pd.DataFrame:
    records: List[Dict[str, object]] = []
    for site in SITES:
        baseline_name = site_best_baseline[site]
        main_run = find_run_dir(site, PROPOSED_MODEL)
        base_run = find_run_dir(site, baseline_name)
        pred_main, label_main = load_event_series(main_run)
        pred_base, label_base = load_event_series(base_run)
        n_events = min(len(pred_main), len(pred_base), len(label_main), len(label_base))

        peaks = [float(np.max(np.asarray(label_main[i]).flatten())) for i in range(n_events)]
        best_idx = int(np.argmax(peaks))
        y_true = np.asarray(label_main[best_idx]).flatten()
        y_main = np.asarray(pred_main[best_idx]).flatten()
        y_base = np.asarray(pred_base[best_idx]).flatten()
        m_main = compute_overall_metrics(y_true, y_main)
        m_base = compute_overall_metrics(y_true, y_base)

        rec: Dict[str, object] = {
            "Site": site_tag(site),
            "CaseID": f"{site_tag(site)}_E{best_idx+1}",
            "EventIndex_1based": best_idx + 1,
            "Duration_min": len(y_true),
            "PeakRain_mmph": float(np.max(y_true)),
            "BestBaseline": model_tag(baseline_name),
            "RMSE_Proposed": m_main["RMSE"],
            "RMSE_Baseline": m_base["RMSE"],
            "Delta_RMSE_pct": (m_base["RMSE"] - m_main["RMSE"]) / m_base["RMSE"] * 100.0,
            "MAE_Proposed": m_main["MAE"],
            "MAE_Baseline": m_base["MAE"],
            "NSE_Proposed": m_main["NSE"],
            "NSE_Baseline": m_base["NSE"],
        }
        records.append(rec)
    return pd.DataFrame(records)


def build_4_2_table(site_dfs: Dict[str, pd.DataFrame], site_best_baseline: Dict[str, str]) -> pd.DataFrame:
    rows: List[pd.DataFrame] = []
    for site, df in site_dfs.items():
        best_rmse = float(df[df["Model"] == site_best_baseline[site]]["RMSE"].iloc[0])
        tmp = df.copy()
        tmp["Site"] = site_tag(site)
        tmp["Model"] = tmp["Model"].map(model_tag)
        tmp["BestBaseline"] = model_tag(site_best_baseline[site])
        tmp["Delta_RMSE_vs_BestBaseline_pct"] = (best_rmse - tmp["RMSE"]) / best_rmse * 100.0
        rows.append(tmp)
    out = pd.concat(rows, ignore_index=True)
    out = out.sort_values(["Site", "RMSE"], ascending=[True, True]).reset_index(drop=True)
    return out


def build_4_5_table(site_ablation_dfs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: List[pd.DataFrame] = []
    for site, df in site_ablation_dfs.items():
        main_rmse = float(df[df["Model"] == PROPOSED_MODEL]["RMSE"].iloc[0])
        tmp = df.copy()
        tmp["Site"] = site_tag(site)
        tmp["Model"] = tmp["Model"].map(model_tag)
        tmp["Delta_RMSE_vs_Proposed_pct"] = (tmp["RMSE"] - main_rmse) / main_rmse * 100.0
        rows.append(tmp)
    out = pd.concat(rows, ignore_index=True)
    out = out.sort_values(["Site", "Delta_RMSE_vs_Proposed_pct"]).reset_index(drop=True)
    return out


def _concat_event_pair(site: str, model_name: str) -> Tuple[np.ndarray, np.ndarray]:
    run_dir = find_run_dir(site, model_name)
    pred_list, label_list = load_event_series(run_dir)
    y_true_all: List[np.ndarray] = []
    y_pred_all: List[np.ndarray] = []
    n_events = min(len(pred_list), len(label_list))
    for i in range(n_events):
        y_true = np.asarray(label_list[i]).flatten()
        y_pred = np.asarray(pred_list[i]).flatten()
        n = min(len(y_true), len(y_pred))
        if n <= 0:
            continue
        y_true_all.append(y_true[:n])
        y_pred_all.append(y_pred[:n])
    return np.concatenate(y_true_all), np.concatenate(y_pred_all)


def _bootstrap_delta_rmse_ci(
    y_true: np.ndarray,
    y_main: np.ndarray,
    y_base: np.ndarray,
    n_boot: int = 1200,
    seed: int = 2026,
) -> Tuple[float, float]:
    n = len(y_true)
    if n <= 1:
        return np.nan, np.nan
    rng = np.random.default_rng(seed)
    vals = np.empty(n_boot, dtype=float)
    for i in range(n_boot):
        idx = rng.integers(0, n, size=n)
        rmse_main_b = rmse(y_true[idx], y_main[idx])
        rmse_base_b = rmse(y_true[idx], y_base[idx])
        vals[i] = (rmse_base_b - rmse_main_b) / rmse_base_b * 100.0 if rmse_base_b > 0 else np.nan
    vals = vals[~np.isnan(vals)]
    if len(vals) == 0:
        return np.nan, np.nan
    return float(np.quantile(vals, 0.025)), float(np.quantile(vals, 0.975))


def build_4_6_significance_table(site_best_baseline: Dict[str, str]) -> pd.DataFrame:
    rows: List[Dict[str, object]] = []
    for site in SITES:
        baseline = site_best_baseline[site]
        y_true_main, y_main = _concat_event_pair(site, PROPOSED_MODEL)
        y_true_base, y_base = _concat_event_pair(site, baseline)
        n = min(len(y_true_main), len(y_main), len(y_true_base), len(y_base))
        y_true = y_true_main[:n]
        y_main = y_main[:n]
        y_base = y_base[:n]

        rmse_main = rmse(y_true, y_main)
        rmse_base = rmse(y_true, y_base)
        delta_rmse_pct = (rmse_base - rmse_main) / rmse_base * 100.0 if rmse_base > 0 else np.nan
        ci_low, ci_high = _bootstrap_delta_rmse_ci(y_true, y_main, y_base)

        abs_err_main = np.abs(y_main - y_true)
        abs_err_base = np.abs(y_base - y_true)

        try:
            p_t = float(stats.ttest_rel(abs_err_base, abs_err_main, alternative="greater").pvalue)
        except Exception:
            p_t = np.nan

        try:
            p_w = float(stats.wilcoxon(abs_err_base, abs_err_main, alternative="greater").pvalue)
        except Exception:
            p_w = np.nan

        rows.append(
            {
                "Site": site_tag(site),
                "BestBaseline": model_tag(baseline),
                "N_points": int(n),
                "RMSE_Proposed": rmse_main,
                "RMSE_Baseline": rmse_base,
                "Delta_RMSE_pct": delta_rmse_pct,
                "Bootstrap95CI_Low": ci_low,
                "Bootstrap95CI_High": ci_high,
                "Paired_t_pvalue": p_t,
                "Wilcoxon_pvalue": p_w,
            }
        )

    return pd.DataFrame(rows)


def save_tables_excel(
    table_42: pd.DataFrame,
    table_43: pd.DataFrame,
    table_44: pd.DataFrame,
    table_45: pd.DataFrame,
    table_46: pd.DataFrame,
) -> Path:
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    out_xlsx = FIG_DIR / "section_4_2_4_5_tables.xlsx"
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        table_42.to_excel(writer, sheet_name="4.2_Overall", index=False)
        table_43.to_excel(writer, sheet_name="4.3_Scenario_Overall", index=False)
        table_44.to_excel(writer, sheet_name="4.4_Case_Metadata", index=False)
        table_45.to_excel(writer, sheet_name="4.5_Ablation", index=False)
        table_46.to_excel(writer, sheet_name="4.6_Significance", index=False)
    return out_xlsx


def append_table_1_2_3_excel(
    table_1: pd.DataFrame,
    table_2: pd.DataFrame,
    table_3: pd.DataFrame,
    out_xlsx: Path | None = None,
) -> Path:
    target = out_xlsx or (FIG_DIR / "section_4_2_4_5_tables.xlsx")
    FIG_DIR.mkdir(parents=True, exist_ok=True)
    mode = "a" if target.exists() else "w"
    writer_kwargs = {"engine": "openpyxl", "mode": mode}
    if mode == "a":
        writer_kwargs["if_sheet_exists"] = "replace"

    with pd.ExcelWriter(target, **writer_kwargs) as writer:
        table_1.to_excel(writer, sheet_name="Table1_SiteInfo", index=False)
        table_2.to_excel(writer, sheet_name="Table2_Variables", index=False)
        table_3.to_excel(writer, sheet_name="Table3_SampleStats", index=False)
    return target


def main() -> None:
    FIG_DIR.mkdir(parents=True, exist_ok=True)

    site_dfs = {site: read_overall_table(site) for site in SITES}
    site_best_baseline = {site: select_best_baseline(df) for site, df in site_dfs.items()}
    site_ablation_dfs = {site: read_ablation_table(site) for site in SITES}

    table_42 = build_4_2_table(site_dfs, site_best_baseline)
    table_43 = build_scenario_table(site_best_baseline)
    table_44 = build_case_table(site_best_baseline)
    table_45 = build_4_5_table(site_ablation_dfs)
    table_46 = build_4_6_significance_table(site_best_baseline)
    table_1 = build_table1_site_info()
    table_2 = build_table2_variable_definition()
    table_3 = build_table3_sample_statistics()

    out_xlsx = save_tables_excel(table_42, table_43, table_44, table_45, table_46)
    append_table_1_2_3_excel(table_1, table_2, table_3, out_xlsx=out_xlsx)

    print(f"[saved] {out_xlsx}")


if __name__ == "__main__":
    main()

